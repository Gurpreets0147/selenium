from openpyxl import Workbook
from faker import Faker
fake_data = Faker()
wb = Workbook()
ws1 = wb.active
# File Name.
dest_filename = 'empty_book.xlsx'
# Add headers
ws1.cell(column=1, row=1, value="Name")
ws1.cell(column=2, row=1, value="Email")
ws1.cell(column=3, row=1, value="Password")
ws1.cell(column=4, row=1, value="Address")
# add random data
for row in range(2, 10):
    ws1.cell(column=1, row=row, value=fake_data.name())
    ws1.cell(column=2, row=row, value=fake_data.email())
    ws1.cell(column=3, row=row, value=fake_data.password())
    ws1.cell(column=4, row=row, value=fake_data.address())

wb.save(filename=dest_filename)
