from openpyxl import Workbook

wb = Workbook()
# File Name.
dest_filename = 'empty_book.xlsx'
# check which sheet is active
ws1 = wb.active
# change name of spreed sheet
ws1.title = "range"
# add data in rows
for row in range(1, 5):
    ws1.append(range(10))
# create new spread sheet
ws2 = wb.create_sheet(title="Pi")
# add data in perticular cell
ws2['A1'] = 3.14
# create new spread sheet
ws3 = wb.create_sheet(title="Data")
# add data in file
for row in range(1, 5):
    for col in range(1, 5):
        ws3.cell(column=col, row=row, value="1")
# print perticular cell value
print(ws3['AA10'].value)
# Save file
wb.save(filename=dest_filename)
